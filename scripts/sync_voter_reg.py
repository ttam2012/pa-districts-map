#!/usr/bin/env python3
"""
Refreshes the county-level voter-registration numbers embedded in
pa-districts-map.html from the PA Dept. of State's official weekly export.

Source: https://www.pa.gov/agencies/dos/resources/voting-and-elections-resources/voting-and-election-statistics
File:   currentvotestats.xlsx -> sheet "Reg Voter"

What this touches:
  - VOTER_REG.cnty in the HTML (the only piece of embedded data that is
    actually county-level voter registration). VOTER_REG.us_house /
    .state_senate / .state_house are a STATIC county-to-district overlap
    map derived from district boundaries, not registration data, and are
    intentionally left untouched.
  - The handful of "as of <Month Year>" citation strings tied directly to
    that data (statewide dossier voter-reg tile + sources footer). Every
    other occurrence of a date string in the file is left alone.

Fails loudly (nonzero exit) rather than guessing if the source's shape
changes -- an unattended weekly job should never silently write bad data.
"""
import datetime
import json
import re
import sys
import urllib.request
from pathlib import Path

XLSX_URL = "https://www.pa.gov/content/dam/copapwp-pagov/en/dos/resources/voting-and-elections/voting-and-election-statistics/currentvotestats.xlsx"
REPO_ROOT = Path(__file__).resolve().parent.parent
HTML_PATH = REPO_ROOT / "pa-districts-map.html"

# The 67 PA counties as currently keyed in VOTER_REG.cnty -- used as a hard
# sanity check. If the source ever adds/removes/renames a county, this
# script stops instead of writing a mismatched dataset.
EXPECTED_COUNTIES = {
    "Adams", "Allegheny", "Armstrong", "Beaver", "Bedford", "Berks", "Blair",
    "Bradford", "Bucks", "Butler", "Cambria", "Cameron", "Carbon", "Centre",
    "Chester", "Clarion", "Clearfield", "Clinton", "Columbia", "Crawford",
    "Cumberland", "Dauphin", "Delaware", "Elk", "Erie", "Fayette", "Forest",
    "Franklin", "Fulton", "Greene", "Huntingdon", "Indiana", "Jefferson",
    "Juniata", "Lackawanna", "Lancaster", "Lawrence", "Lebanon", "Lehigh",
    "Luzerne", "Lycoming", "McKean", "Mercer", "Mifflin", "Monroe",
    "Montgomery", "Montour", "Northampton", "Northumberland", "Perry",
    "Philadelphia", "Pike", "Potter", "Schuylkill", "Snyder", "Somerset",
    "Sullivan", "Susquehanna", "Tioga", "Union", "Venango", "Warren",
    "Washington", "Wayne", "Westmoreland", "Wyoming", "York",
}

# The sheet spells most counties in plain upper case, title-casing cleanly
# -- McKean is the one PA county that doesn't.
NAME_OVERRIDES = {"MCKEAN": "McKean"}


def normalize_county_name(raw: str) -> str:
    raw = raw.strip()
    if raw.upper() in NAME_OVERRIDES:
        return NAME_OVERRIDES[raw.upper()]
    return raw.title()


def fetch_xlsx(dest: Path) -> Path:
    req = urllib.request.Request(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        dest.write_bytes(resp.read())
    return dest


def parse_reg_voter_sheet(xlsx_path: Path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Reg Voter" not in wb.sheetnames:
        sys.exit(f"FATAL: 'Reg Voter' sheet not found. Sheets present: {wb.sheetnames}")
    ws = wb["Reg Voter"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0 carries "Information as of MM/DD/YYYY" in cell A1.
    as_of_cell = str(rows[0][0] or "")
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", as_of_cell)
    if not m:
        sys.exit(f"FATAL: couldn't find an 'as of MM/DD/YYYY' date in row 1: {as_of_cell!r}")
    as_of = datetime.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))

    counties = {}
    # Data starts at row index 2 (0-based) -- row 0 is the "as of" banner,
    # row 1 is the column header.
    for row in rows[2:]:
        name_raw = row[0]
        if not name_raw or not str(name_raw).strip():
            continue
        if str(name_raw).strip().lower().startswith("total"):
            continue  # the sheet's trailing "Totals:" summary row, not a county
        try:
            dem = int(row[2])
            rep = int(row[4])
            no_aff = int(row[6])
            other = int(row[8])
            total = int(row[10])
        except (TypeError, ValueError, IndexError):
            continue  # skip any stray footer/summary row that isn't county data
        name = normalize_county_name(str(name_raw))
        combined_other = no_aff + other
        if dem + rep + combined_other != total:
            sys.exit(
                f"FATAL: {name} figures don't sum to the reported total "
                f"({dem}+{rep}+{combined_other} != {total}) -- source format may have changed."
            )
        counties[name] = {"d": dem, "r": rep, "o": combined_other, "t": total}

    got = set(counties)
    if got != EXPECTED_COUNTIES:
        missing = EXPECTED_COUNTIES - got
        extra = got - EXPECTED_COUNTIES
        sys.exit(
            "FATAL: parsed county set doesn't match the expected 67 PA counties.\n"
            f"  missing: {sorted(missing)}\n  unexpected: {sorted(extra)}"
        )

    return as_of, counties


def build_cnty_json(counties: dict) -> str:
    # Preserve the file's existing alphabetical key order and its exact
    # compact formatting (no spaces) so the diff stays minimal.
    parts = []
    for name in sorted(counties):
        c = counties[name]
        parts.append(f'"{name}":{{"d":{c["d"]},"r":{c["r"]},"o":{c["o"]},"t":{c["t"]}}}')
    return "{" + ",".join(parts) + "}"


def splice_into_html(html: str, cnty_json: str, as_of: datetime.date) -> tuple[str, bool]:
    changed = False

    # 1. VOTER_REG.cnty -- bounded by the literal '"cnty":{' start and the
    #    known-adjacent '},"us_house"' that immediately follows it.
    pattern = re.compile(r'"cnty":\{.*?\},"us_house"', re.DOTALL)
    if not pattern.search(html):
        sys.exit('FATAL: could not locate `"cnty":{...},"us_house"` in pa-districts-map.html -- refusing to guess.')
    new_html, n = pattern.subn(f'"cnty":{cnty_json},"us_house"', html, count=1)
    if n != 1:
        sys.exit("FATAL: VOTER_REG.cnty replacement did not apply exactly once.")
    if new_html != html:
        changed = True
    html = new_html

    # 2. The citation strings tied directly to this data. Each is matched
    #    structurally (a regex anchored to its unique surrounding markup)
    #    rather than against one fixed old value -- so every weekly run
    #    updates them correctly regardless of what date they currently show,
    #    not just the very first run after this script was introduced. An
    #    unrelated date elsewhere in the file (e.g. an official's term) can
    #    never be touched, since the surrounding markup is always matched too.
    label = as_of.strftime("%b %Y")            # e.g. "Aug 2026"
    full_label = as_of.strftime("%B %-d, %Y") if sys.platform != "win32" \
        else as_of.strftime("%B %#d, %Y")      # e.g. "August 31, 2026"
    month_year = r"[A-Za-z]+ \d{4}"
    full_date = r"[A-Za-z]+ \d{1,2}, \d{4}"
    citation_patterns = [
        (rf"<span>Registered voters \({month_year}\)</span>",
         f"<span>Registered voters ({label})</span>"),
        (rf'<div class="stat-sub">{month_year} \(PA DOS\)</div>',
         f'<div class="stat-sub">{label} (PA DOS)</div>'),
        (rf"Sources: PA Dept\. of State \(voter registration {month_year} · 2024 election returns\)",
         f"Sources: PA Dept. of State (voter registration {label} · 2024 election returns)"),
        (rf'(class="h3-updated-link">Last updated: ){full_date}(</a>)',
         rf"\g<1>{full_label}\g<2>"),
    ]
    for pattern_str, new in citation_patterns:
        pattern = re.compile(pattern_str)
        if not pattern.search(html):
            continue  # cosmetic citation text, not the data-integrity-critical part
        # count=0 (unlimited) -- some of these, like the "Last updated" link,
        # now appear in more than one place in the file (statewide dossier,
        # county dossier, per-district voter-reg section) and all of them
        # share the same underlying data snapshot, so all should update.
        new_html, n = pattern.subn(new, html, count=0)
        if new_html != html:
            changed = True
        html = new_html

    return html, changed


def main():
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, help="use a local .xlsx instead of downloading")
    parser.add_argument("--dry-run", action="store_true", help="parse and validate only, don't write the HTML file")
    args = parser.parse_args()

    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        xlsx_path = Path(__file__).resolve().parent / "_currentvotestats.xlsx"
        print(f"Downloading {XLSX_URL} ...")
        fetch_xlsx(xlsx_path)

    as_of, counties = parse_reg_voter_sheet(xlsx_path)
    total_registered = sum(c["t"] for c in counties.values())
    print(f"Parsed {len(counties)} counties, as of {as_of.isoformat()}, "
          f"{total_registered:,} total registered voters.")

    if args.dry_run:
        print("Dry run -- not writing pa-districts-map.html.")
        return

    html = HTML_PATH.read_text(encoding="utf-8")
    cnty_json = build_cnty_json(counties)
    new_html, changed = splice_into_html(html, cnty_json, as_of)

    if not changed:
        print("No changes -- embedded data already matches the source.")
        return

    HTML_PATH.write_text(new_html, encoding="utf-8")
    print(f"Updated {HTML_PATH} with data as of {as_of.isoformat()}.")

    # GitHub Actions: surface whether anything actually changed to the
    # workflow step via GITHUB_OUTPUT, so the commit step can be skipped
    # cleanly on no-op weeks.
    import os
    gh_out = os.environ.get("GITHUB_OUTPUT")
    if gh_out:
        with open(gh_out, "a", encoding="utf-8") as f:
            f.write(f"changed=true\nas_of={as_of.isoformat()}\ntotal={total_registered}\n")


if __name__ == "__main__":
    main()
